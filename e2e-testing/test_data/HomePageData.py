import openpyxl

class HomePageData:
  # test_homepage_data = [
  #           {"email": "hello@hello.com", "password": "123456", "name": "hello"},
  #           {"email": "cmj@cmj.com", "password": "qwer1234", "name": "hi"},
  #       ]
  
  @staticmethod
  def getTestData(test_case_name):
    excel_file = openpyxl.load_workbook("/Users/chaminjae/Desktop/SQA/python_excel_demo.xlsx")
    # 시트 제어
    sheet = excel_file.active
    dict = {}
    for i in range(1, sheet.max_row+1):
      for j in range(sheet.max_column):
          if sheet[i][0].value == "Testcase2":
              #딕셔너리에 넣기
              dict[sheet[1][j].value] = sheet[i][j].value
    
    return [dict]