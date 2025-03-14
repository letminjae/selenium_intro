import openpyxl

excel_file = openpyxl.load_workbook("/Users/chaminjae/Desktop/SQA/python_excel_demo.xlsx")

# 시트 제어
sheet = excel_file.active
dict = {}

# 시트 값 확인하는 방법
cell = sheet.cell(row=1, column=2) # 성
B1 = sheet['B1'].value # 성

# cell 값을 변경함 - 실제 엑셀에선 아직 변경되지 않음.
sheet.cell(row=2, column=2).value = "Cha" 
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row) #전체 행
print(sheet.max_column) # 전체 column 열

# 엑셀 시트 값 출력
for i in range(1, sheet.max_row+1):
    for j in range(sheet.max_column):
        if sheet[i][0].value == "Testcase2":
            #딕셔너리에 넣기
            dict[sheet[1][j].value] = sheet[i][j].value
    
print(dict)